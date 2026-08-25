"""按公司把发票记录汇总到一个 Excel 文件，月份做 sheet。
不再依赖 raw/ 或 invoices/，只接收已解析的 dict 列表。

使用：
    python build_excel.py --config config.json
    python build_excel.py --config config.json --dry-run
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "序号", "开票日期", "发票号码", "项目名称",
    "数量", "单价", "金额(不含税)", "税率", "税额",
    "价税合计(小写)", "价税合计(大写)",
    "销售方", "销售方税号", "购买方", "购买方税号",
    "备注",
]
SUM_COLS = {"金额(不含税)", "税额", "价税合计(小写)"}
SUMMARY_HEADERS = ["月份", "发票张数", "金额(不含税)合计", "税额合计", "价税合计(小写)合计", "备注"]


def match_company(buyer: str, companies: list[dict]) -> str | None:
    for c in companies:
        if c["match"] in buyer:
            return c["key"]
    return None


def to_number(s) -> float | str:
    if isinstance(s, (int, float)):
        return s
    try:
        return float(s)
    except (TypeError, ValueError):
        return s if s else ""


def collect_all(manifest_rows: list[dict], months: list[str]) -> list[dict]:
    return [r for r in manifest_rows if r.get("月份") in months]


def _apply_widths(ws, widths: list[int]) -> None:
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _style_header(ws, title: str, col_count: int, headers: list[str] | None = None) -> None:
    if headers is None:
        headers = HEADERS
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    c = ws.cell(row=1, column=1)
    c.font = Font(name="Microsoft YaHei", size=14, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="305496")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    thin = Side(border_style="thin", color="9E9E9E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 30


def _format_month_sheet(ws, company_key: str, month: str, rows: list[dict]) -> None:
    title = f"{rows[0].get('购买方', company_key)} — {month[:4]}年{int(month[5:])}月 开票记录"
    _style_header(ws, title, len(HEADERS))

    thin = Side(border_style="thin", color="9E9E9E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(name="Microsoft YaHei", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    for i, r in enumerate(rows, start=1):
        excel_row = i + 2
        values = [
            i, r.get("开票日期", ""), r.get("发票号码", ""), r.get("项目名称", ""),
            to_number(r.get("数量", "")), to_number(r.get("单价", "")),
            to_number(r.get("金额(不含税)", "")), r.get("税率", ""),
            to_number(r.get("税额", "")), to_number(r.get("价税合计(小写)", "")),
            r.get("价税合计(大写)", ""), r.get("销售方", ""), r.get("销售方税号", ""),
            r.get("购买方", ""), r.get("购买方税号", ""), r.get("备注", ""),
        ]
        for j, v in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=j, value=v)
            cell.font = body_font
            cell.border = border
            h = HEADERS[j - 1]
            if h in ("序号", "数量", "税率"):
                cell.alignment = center
            elif h in ("金额(不含税)", "单价", "税额", "价税合计(小写)"):
                cell.alignment = right
                if isinstance(v, (int, float)):
                    cell.number_format = "#,##0.00"
            else:
                cell.alignment = left

    # 合计行
    last_data_row = len(rows) + 2
    total_row = last_data_row + 1
    ws.cell(row=total_row, column=1, value="合计")
    for j, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=total_row, column=j)
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
        cell.border = border
        if h in SUM_COLS:
            col = get_column_letter(j)
            cell.value = f"=SUM({col}3:{col}{last_data_row})"
            cell.number_format = "#,##0.00"
            cell.alignment = right
        elif h == "项目名称":
            cell.value = f"共 {len(rows)} 张发票"
            cell.alignment = center
        else:
            cell.alignment = center

    _apply_widths(ws, [6, 14, 22, 26, 6, 12, 14, 8, 12, 16, 22, 22, 22, 22, 22, 30])
    ws.freeze_panes = "A3"


def _format_summary_sheet(ws, company_key: str, months_data: list[tuple[str, list[dict]]]) -> None:
    buyer_full = months_data[0][1][0].get("购买方", company_key) if months_data else company_key
    _style_header(ws, f"{buyer_full} — 跨月汇总", len(SUMMARY_HEADERS), SUMMARY_HEADERS)

    thin = Side(border_style="thin", color="9E9E9E")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    total_n, total_amount, total_tax, total_grand = 0, 0.0, 0.0, 0.0

    for i, (m, rows) in enumerate(months_data, start=1):
        n = len(rows)
        s_amount = sum(float(r.get("金额(不含税)", 0) or 0) for r in rows)
        s_tax = sum(float(r.get("税额", 0) or 0) for r in rows)
        s_grand = sum(float(r.get("价税合计(小写)", 0) or 0) for r in rows)
        total_n += n
        total_amount += s_amount
        total_tax += s_tax
        total_grand += s_grand

        excel_row = i + 2
        cell = ws.cell(row=excel_row, column=1, value=m)
        cell.hyperlink = f"#'{m}'!A1"
        cell.font = Font(name="Microsoft YaHei", size=10, color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

        for k, v in enumerate([n, s_amount, s_tax, s_grand], start=2):
            fmt = "#,##0.00" if k > 2 else "0"
            cc = ws.cell(row=excel_row, column=k, value=v)
            cc.font = Font(name="Microsoft YaHei", size=10)
            cc.alignment = Alignment(horizontal="right", vertical="center")
            cc.number_format = fmt
            cc.border = border

        notes = "; ".join(sorted({r.get("销售方", "") for r in rows if r.get("销售方")}))
        nc = ws.cell(row=excel_row, column=len(SUMMARY_HEADERS), value=notes or "")
        nc.font = Font(name="Microsoft YaHei", size=10)
        nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        nc.border = border

    # 全期合计行
    total_row = len(months_data) + 3
    ws.cell(row=total_row, column=1, value="全期合计")
    for k, v in enumerate([total_n, total_amount, total_tax, total_grand], start=2):
        fmt = "#,##0.00" if k > 2 else "0"
        cc = ws.cell(row=total_row, column=k, value=v)
        cc.font = Font(name="Microsoft YaHei", size=11, bold=True)
        cc.fill = PatternFill("solid", fgColor="FCE4D6")
        cc.alignment = Alignment(horizontal="right", vertical="center")
        cc.number_format = fmt
        cc.border = border
    for j in range(1, len(SUMMARY_HEADERS) + 1):
        cc = ws.cell(row=total_row, column=j)
        cc.font = Font(name="Microsoft YaHei", size=11, bold=True)
        cc.fill = PatternFill("solid", fgColor="FCE4D6")
        cc.border = border

    _apply_widths(ws, [12, 12, 18, 18, 22, 30])
    ws.freeze_panes = "A3"


def write_company_workbook(
    company_key: str,
    months_data: list[tuple[str, list[dict]]],
    project: Path,
    records_subdir: str,
    dry_run: bool,
) -> Path | None:
    """months_data: [(month, rows), ...] 按月份升序。"""
    if not months_data:
        return None

    out_xlsx = project / records_subdir / f"{company_key}.xlsx"
    if dry_run:
        total_n = sum(len(rows) for _, rows in months_data)
        print(f"  [dry] {out_xlsx.relative_to(project)} ({len(months_data)} 个月, {total_n} 条)")
        return None

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "汇总"
    _format_summary_sheet(ws_summary, company_key, months_data)

    for month, rows in months_data:
        ws = wb.create_sheet(title=month)
        _format_month_sheet(ws, company_key, month, rows)

    wb.save(out_xlsx)
    total_n = sum(len(rows) for _, rows in months_data)
    print(f"  ✓ {out_xlsx.relative_to(project)} ({len(months_data)} 个月, {total_n} 张)")
    return out_xlsx


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", required=True)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    cfg = json.loads(Path(args.config).read_text(encoding="utf-8"))
    project = Path(cfg["project"]).expanduser().resolve()
    records_subdir = cfg.get("records_subdir", "开票记录")
    months = cfg["months"]
    companies = cfg["companies"]

    print("注意：本脚本只从 manifest.json 读取数据，需先运行 ingest.py")
    return 0


if __name__ == "__main__":
    sys.exit(main())
